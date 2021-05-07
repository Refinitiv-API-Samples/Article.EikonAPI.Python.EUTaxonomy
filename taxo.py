# Taxonomy report generation
import eikon as ek
import pandas as pd
import time
from argparse import ArgumentParser
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, NamedStyle, PatternFill

# global fields
TRBC_db = None
TAXON_db = None
TESTING_MET_db = None
ESG_FIELDS = None


#==============================================
def init(appkey):
#==============================================
	# connect to eikon
	ek.set_app_key(appkey)



#==============================================
# load the input portfolio to be analized
def loadInputPortfolio(pFilename):
#==============================================
	workbook = load_workbook(filename = pFilename)

	sheet = workbook.active
	data = sheet.values
	cols = next(data)
	data = list(data)

	inputdf = pd.DataFrame(data, columns=cols)
	inputdf.dropna(inplace=True)

	inputlist = inputdf['RIC'].to_list()
	inputlist = list(filter(None, inputlist))

	return inputlist


#==============================================
# Load the database from excel spreadsheet
def loadDatabase(dbFileName):
#==============================================
	global TRBC_db, TAXON_db, TESTING_MET_db, ESG_FIELDS
	workbook = load_workbook(filename = dbFileName)

	sheet = workbook['NAICS>TRBC']
	data = sheet.values
	cols = next(data)
	data = list(data)
	TRBC_db = pd.DataFrame(data, columns=cols)

	sheet = workbook['EU Taxonomy']
	data = sheet.values
	cols = next(data)
	data = list(data)
	TAXON_db = pd.DataFrame(data, columns=cols)

	sheet = workbook['Testing Metrics']
	data = sheet.values
	cols = next(data)
	data = list(data)
	TESTING_MET_db = pd.DataFrame(data, columns=cols)

	# get all the unique ESG fields from the database
	ESG_FIELDS = list(set(TESTING_MET_db['Refinitiv ESG Field'].to_list()))
	if None in ESG_FIELDS:
		ESG_FIELDS.remove(None)

	ESG_FIELDS.append('TR.CommonName')
	ESG_FIELDS.append('TR.TRESGScore')
	ESG_FIELDS.append('TR.TRBCActivityCode')
	ESG_FIELDS.append('TR.TRBCEconomicSector')
	ESG_FIELDS.append('TR.TRBCActivity')



#==============================================
# Get the taxonomy data for a RICs from Eikon/RDP
def getData(inputlist):
#==============================================
	# get sector revenue data
	txk, err = ek.get_data(inputlist, ['TR.BGS.BusTotalRevenue.segmentCode', 'TR.BGS.BusTotalRevenue.segmentName', 'TR.BGS.BusTotalRevenue.fperiod', 'TR.BGS.BusTotalRevenue.currency', 'TR.BGS.BusTotalRevenue.value'])
	# also get the ESG data for all the fields
	esg_f, err = ek.get_data(inputlist, ESG_FIELDS)
	dnsh, err = ek.get_data(inputlist, ['TR.ControvEnv','TR.RecentControvEnv','TR.ControvCopyrights','TR.ControvPublicHealth','TR.ControvBusinessEthics','TR.ControvTaxFraud','TR.ControvAntiCompetition','TR.ControvCriticalCountries','TR.RecentControvPublicHealth','TR.RecentControvBusinessEthics','TR.RecentControvTaxFraud','TR.RecentControvAntiCompetition','TR.RecentControvCriticalCountries','TR.RecentControvCopyrights','TR.ControvHumanRights','TR.ControvChildLabor','TR.RecentControvHumanRights','TR.RecentControvChildLabor','TR.ControvConsumer','TR.RecentControvConsumer','TR.ControvCustomerHS','TR.ControvResponsibleRD','TR.ControvPrivacy','TR.ControvRespMarketing','TR.ControvProductAccess','TR.RecentControvCustomerHS','TR.RecentControvPrivacy','TR.RecentControvRespMarketing','TR.RecentControvProductAccess','TR.RecentControvResponsibleRD','TR.Strikes','TR.ControvEmployeesHS','TR.RecentControvEmployeesHS','TR.EnvProducts','TR.LandEnvImpactReduction','TR.EcoDesignProducts'])

	return txk, esg_f, dnsh



#==============================================
# handle case when no data is available for a RIC
def processEmpty(ric, buisData, esgData):
#==============================================
	aggD = {
		'Instrument': ric,
		'Name': esgData['Company Common Name'][0],
		'Delisted': 'Yes' if '^' in ric else '',
		'ESG Score': esgData['ESG Score'][0],
		'Economic Sector': esgData['TRBC Economic Sector Name'][0],
		'TRBC Activity': esgData['TRBC Activity Name'][0]
	}

	# Step 15: Is parent company eligible (if trbc data is not available)
	#-----------------------------------
	parentTRBCode = esgData['TRBC Activity Code'][0]
	if not pd.isnull(parentTRBCode):
		parentTxnMatch = TAXON_db[TAXON_db['TRBC code'] == parentTRBCode]
		if parentTxnMatch.empty:
			aggD['Parent Eligible'] = 'Not in scope'
			aggD['Parent Not In Scope ratio'] = 1
		else:
			aggD['Parent Eligible'] = parentTxnMatch.iloc[0]['Additional testing needed?']
			aggD['Parent Eligible ratio'] = 1

	buisData['Name'] = esgData['Company Common Name'][0]
	buisData['Delisted'] = 'Delisted, No Data' if '^' in ric else 'No Data'
	return aggD, buisData



#==============================================
# process taxonomy data for single RIC
def getTaxoForRic(ric, buisData, esgData):
#==============================================
	if pd.isnull(buisData['Business Total Revenues (Calculated)'][0]):
		return processEmpty(ric, buisData, esgData)

	# Step 3: Calculate the segment revenue share
	txkSeg = pd.DataFrame(buisData[~buisData['Segment Code'].str.match('SEGMTL|ICELIM|EXPOTH|CONSTL')])
	revList = txkSeg['Business Total Revenues (Calculated)'].to_list()
	if sum(revList) < 10:
		return processEmpty(ric, buisData, esgData)
		
	segRevenueRatio = [x/sum(revList) for x in revList]
	#print('Segment revenue ratio: %s' % segRevenueRatio)

	# process all the segment codes
	txkSeg.insert(1, 'Name', esgData['Company Common Name'][0])
	txkSeg.insert(2, 'Delisted', 'Delisted' if '^' in ric else '')
	txkSeg['Segment Revenue Ratio'] = segRevenueRatio
	txkSeg['TRBC Codes'] = ''
	txkSeg['Match with EU Taxo'] = ''
	txkSeg['Linked Assesment Metric'] = ''
	txkSeg['Metric Reported Value'] = ''
	txkSeg['Threshold Test'] = ''
	txkSeg['Segment Weight'] = ''

	txkSeg['Aligned'] = 0.
	txkSeg['Additional Testing Required'] = 0.
	txkSeg['Not in Scope'] = 0.
	txkSeg['Others'] = 0.
	txkSeg['Aligned- Pass'] = 0.
	txkSeg['Aligned- No Data'] = 0.
	txkSeg['Aligned- Not in Scope'] = 0.

	for idx in range(len(txkSeg)):

		# Step 5: Convert NAICS code to TRBC codes for every segment
		#-----------------------------------
		segCodeList = txkSeg.iloc[idx]['Segment Code'].split(',')
		trbcCodeList = []
		for naicCode in segCodeList:
			if(naicCode.isnumeric()):
				# append 0 if the code is < 6 chars
				if len(naicCode) < 6:
					naicCode = naicCode + '0'
				# lookup the NIACS -> TRBC code
				trbMatch = TRBC_db[TRBC_db['NAICS Code'] == int(naicCode)]
				if trbMatch.empty:
					trbcCodeList.append(0)
				else:
					trbcCodeList.append(trbMatch.iloc[0]['TRBC Hierarchical Code'])

		#print('%i: NAICS: %s, TRBC: %s' % (idx, segCodeList, trbcCodeList))
		txkSeg.at[idx, 'TRBC Codes'] = ', '.join(str(e) for e in trbcCodeList)


		# Step 6: match against EU taxonomy
		#-----------------------------------
		matchAgainstTaxo = []
		for tCode in trbcCodeList:
			txnMatch = TAXON_db[TAXON_db['TRBC code'] == tCode]
			if txnMatch.empty:
				matchAgainstTaxo.append('na')
			else:
				matchAgainstTaxo.append(txnMatch.iloc[0]['Additional testing needed?'])

		#print('%i: Matching with EU Taxonomy: %s' % (idx, matchAgainstTaxo))
		txkSeg.at[idx, 'Match with EU Taxo'] = ', '.join(str(e) for e in matchAgainstTaxo)


		# Step 7: Is TRBC Code aligned to assesement metric
		#-----------------------------------
		alignedMetricName = []
		alignedMetricField = []
		for tCode in trbcCodeList:
			metMatch = TESTING_MET_db[TESTING_MET_db['TRBC Activity'] == tCode]
			if not metMatch.empty:
				alignedMetricName.append(metMatch.iloc[0]['Refinitiv ESG Data Measures'])
				alignedMetricField.append(metMatch.iloc[0]['Refinitiv ESG Field'])
			else:
				alignedMetricName.append('')

		#print('Is business segment aligned: %s' % alignedMetricName)
		if not all('' == s for s in alignedMetricName):
			txkSeg.at[idx, 'Linked Assesment Metric'] = ', '.join(str(e) for e in alignedMetricName)

		alignedMetricField = list(set(alignedMetricField))
		#print('ESG fields used: %s' % alignedMetricField)


		# Step 8: What is company reported value for aligned metric
		#-----------------------------------
		if len(alignedMetricField) > 0:
			repValues = []
			for almn in alignedMetricName:
				if almn in esgData.columns:
					repValues.append(esgData[almn][0])
				else:
					repValues.append('')

			#print('Reported values: %s' % repValues)
			#if not all(pd.isna(s) or '' == s for s in repValues):
			txkSeg.at[idx, 'Metric Reported Value'] = ', '.join(str(e) for e in repValues)

			# Step 9: Does it pass threashold test
			#-----------------------------------
			threasoldValues = []
			for tCode in trbcCodeList:
				metMatch = TESTING_MET_db[TESTING_MET_db['TRBC Activity'] == tCode]
				if metMatch.empty:
					threasoldValues.append('')
				else:
					threasoldValues.append(metMatch.iloc[0]['Used for testing'])

			#print('Threshold values: %s' % threasoldValues)
			thresholdTest = []
			for i in range(len(repValues)):
				if pd.isnull(repValues[i]):
					thresholdTest.append('Data not available')
				elif not repValues[i]:
					thresholdTest.append('')
				elif repValues[i] > threasoldValues[i]:
					thresholdTest.append('Not in Scope')
				else:
					thresholdTest.append('Pass - Aligned')

			#print('Threshold test: %s' % thresholdTest)
			txkSeg.at[idx, 'Threshold Test'] = ', '.join(str(e) for e in thresholdTest)
			# Step 12: Count the TRBC which have passed/failed/require more testing on the threshold
			#print('Count of codes passed/failed etc: %s' % {i:thresholdTest.count(i) for i in thresholdTest})


		# Step 10: What is the weight of each code per segment
		#-----------------------------------
		weightOfEachCode = 1/len(segCodeList)
		#print('Weight of each TRBC code per segment: %s' % weightOfEachCode)
		txkSeg.at[idx, 'Segment Weight'] = weightOfEachCode


		# Step 13: Convert taxo result into %
		#-----------------------------------
		segRev = segRevenueRatio[idx] * weightOfEachCode
		txkSeg.at[idx, 'Aligned'] = segRev * matchAgainstTaxo.count('No')
		txkSeg.at[idx, 'Additional Testing Required'] = segRev * matchAgainstTaxo.count('Yes')
		txkSeg.at[idx, 'Not in Scope'] = segRev * matchAgainstTaxo.count('na')

		if len(trbcCodeList) == 0:
			txkSeg.at[idx, 'Others'] = segRev

		if len(alignedMetricField) > 0:
			txkSeg.at[idx, 'Aligned- Pass'] = segRev * thresholdTest.count('Pass - Aligned')
			txkSeg.at[idx, 'Aligned- No Data'] = segRev * thresholdTest.count('Data not available')
			txkSeg.at[idx, 'Aligned- Not in Scope'] = segRev * thresholdTest.count('Not in Scope')


	# Step 14: Aggregate the business segments into parent company
	#-----------------------------------
	sumSeries = txkSeg.sum(axis=0)
	aggD = {
		'Instrument': ric,
		'Name': esgData['Company Common Name'][0],
		'Delisted': 'Yes' if '^' in ric else '',
		'ESG Score': esgData['ESG Score'][0],
		'Economic Sector': esgData['TRBC Economic Sector Name'][0],
		'TRBC Activity': esgData['TRBC Activity Name'][0],
		'Aligned by Industry': sumSeries['Aligned'],
		'Additional Testing Required': sumSeries['Additional Testing Required'],
		'Eligible': sumSeries['Aligned'] + sumSeries['Additional Testing Required'],
		'Not In Scope': sumSeries['Not in Scope'],
		'Others': sumSeries['Others'],
		'Aligned- Pass': sumSeries['Aligned- Pass'],
		'Aligned- No Data': sumSeries['Aligned- No Data'],
		'Aligned- Not in Scope': sumSeries['Aligned- Not in Scope'],
		'Additional testing needed': sumSeries['Additional Testing Required'] - sumSeries['Aligned- Pass'] + sumSeries['Aligned- Not in Scope'],
		'Total': sumSeries['Aligned'] + sumSeries['Additional Testing Required'] + sumSeries['Not in Scope'] + sumSeries['Others']
	}

	return aggD, txkSeg




#==============================================
def generateReport(rFileName, orgDF, sectorDF, dnshDF):
#==============================================
	sectorDF = sectorDF.astype(object)
	sectorDF.fillna('', inplace=True)
	orgDF.fillna('', inplace=True)
	dnshDF = dnshDF.astype(object)
	dnshTemp = pd.DataFrame(dnshDF)
	dnshDF.fillna('', inplace=True)
	dnshTemp.fillna(0, inplace=True)

	# define the spreadsheet styles
	header = NamedStyle(name="header")
	header.font = Font(bold=True)
	header.border = Border(bottom=Side(border_style="thin"))
	header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


	def addDataFrame(theSht, sourceDF):
		curRow = theSht._current_row
		for row in dataframe_to_rows(sourceDF, index=False, header=True):
			theSht.append(row)
		header_row = theSht[curRow + 1]
		for cell in header_row:
			cell.style = header


	def columnFormats(sheet, fDict):
		for ckey, cformat in fDict.items():
			col = sheet[ckey]
			for cell in col:
				cell.number_format = cformat


	def columnWidths(sheet, sDict):
		for ckey, cWidth in sDict.items():
			sheet.column_dimensions[ckey].width = cWidth


	workbook = Workbook()
	# First sheet with summary data
	ws1 = workbook.active
	ws1.title = 'Organization Summary'
	# Append summary data to first sheet
	summDF = orgDF[['Instrument', 'Name', 'Delisted', 'ESG Score', 'Economic Sector', 'TRBC Activity', 'Eligible', 'Not In Scope', 'Parent Eligible ratio', 'Parent Not In Scope ratio', 'Aligned by Industry', 'Aligned- Pass', 'Additional testing needed', 'Aligned- Not in Scope', 'Others']].copy()
	sum_column = summDF["Aligned by Industry"] + summDF["Aligned- Pass"]
	summDF.insert(11, 'Aligned Total', sum_column)
	sum_column = summDF["Not In Scope"] + summDF["Aligned Total"] + summDF["Additional testing needed"] + summDF["Aligned- Not in Scope"] + summDF["Others"]
	summDF['Total Revenues'] = sum_column
	summDF.columns = ['Instrument', 'Name', 'Delisted?', 'ESG Score', 'TRBC Sector', 'TRBC Activity', '% of business segment revenues eligible', '% of business segment revenues not in scope', 'IF no business segment data is available is the company eligible?',
					  'IF no business segment data is available is the company not in scope?', 'Aligned - By industry activity', 
					  'Aligned - Passed Screening Criteria Threshold Test', 'Aligned Total', 'Additional testing needed', 'Eligible but not aligned (Did not pass threshold test)', 
					  '% FROM OTHER REVENUES', 'Total Revenues %']
	# Merge DNSH data
	contvSum = dnshTemp['Environmental Controversies Count'] + dnshTemp['Recent Environmental Controversies']
	contvSum[contvSum == 0] = pd.NA
	summDF['DNSH Principle - Environment Controversies Count'] = contvSum
	promoteSum = (dnshTemp['Environmental Products'] == 'True') | (dnshTemp['Land Environmental Impact Reduction'] == 'True') | (dnshTemp['Eco-Design Products'] == 'True')
	summDF['Does the company promote environmentally friendly or eco-design products or land impact reduction?'] = promoteSum
	summDF['DNSH - Environment Red Flag (Count > 0 and promotes environmentally products)'] = promoteSum & (contvSum > 0)
	summDF['Minimum Social Safeguards - Social Controversies Count'] = dnshTemp.sum(axis=1)

	summDF.replace({'Does the company promote environmentally friendly or eco-design products or land impact reduction?': {True: 'Yes', False: ''}}, inplace=True)
	summDF.replace({'DNSH - Environment Red Flag (Count > 0 and promotes environmentally products)': {True: 'Flag', False: ''}}, inplace=True)
	summDF.replace({'Minimum Social Safeguards - Social Controversies Count': {0: ''}}, inplace=True)

	addDataFrame(ws1, summDF)
	columnFormats(ws1, {'D': '0.00', 'G': '0%', 'H': '0%', 'I': '0%', 'J': '0%', 'K': '0%', 'L': '0%', 'M': '0%', 'N': '0%', 'O': '0%', 'P': '0%', 'Q': '0%'})
	columnWidths(ws1, {'A': 12, 'B': 40, 'C': 10, 'D': 12, 'E': 25, 'F': 31, 'G': 12, 'H': 12, 'I': 14, 'J': 14, 'K': 12, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 12, 'R': 11, 'S': 14, 'T': 14, 'U': 12})
	ws1.freeze_panes = "B2"

	# Second sheet with sector vise breakdown
	ws3 = workbook.create_sheet("Segment Data Analysis")
	sectorDF.rename(columns={'Segment Code':'NAICS 2007 code',
							 'Delisted': 'Status',
							 'Segment Revenue Ratio': 'Segment Revenue as a %',
							 'Match with EU Taxo':'EU Taxonomy Eligibility\n(Yes: Eligible - Further Testing Needed)\n(No: Eligible - No Further Testing Needed)\n(na: Not in Scope)', 
							 'Linked Assesment Metric':'Technical Screening Criteria test metric',
							 'Threshold Test':'Technical Screening Criteria test results\n(Not in scope: Eligible but not aligned; did not pass threshold test)\n(Data not available: Further testing needed; data not available)\n(Pass Aligned: Aligned and passed technical screening criteria threshold test)',
							 'Aligned':'Percentage Aligned – by industry activity',
							 'Others':'Percentage Other Revenues',
							 'Aligned- Pass':'Percentage Aligned - passed technical screening criteria threshold test',
							 'Aligned- No Data':'Percentage Further testing needed; data not available',
							 'Aligned- Not in Scope':'Percentage Eligible but not aligned (Did not pass threshold test)'
							}, inplace=True)
	addDataFrame(ws3, sectorDF)
	columnFormats(ws3, {'H': '#,##0', 'I': '0%', 'N': '0%', 'O': '0%', 'P': '0%', 'Q': '0%', 'R': '0%', 'S': '0%', 'T': '0%', 'U': '0%', 'V': '0%'})
	columnWidths(ws3, {'A': 12, 'B': 40, 'C': 10, 'D': 30, 'E': 40, 'F': 10, 'G': 10, 'H': 15, 'I': 10, 'J': 46, 'K': 24, 'L': 45, 'M': 20, 'N': 40, 'O': 10, 'P': 12, 'Q': 11, 'R': 10, 'S': 10, 'T': 12, 'U': 12, 'V': 12})
	ws3.freeze_panes = "B2"

	# Third sheet with DNSH data
	ws4 = workbook.create_sheet("DNSH data")
	dnshDF.replace({'Strikes': {'False': ''}}, inplace=True)
	dnshDF.replace({'Environmental Products': {'False': ''}}, inplace=True)
	dnshDF.replace({'Land Environmental Impact Reduction': {'False': ''}}, inplace=True)
	dnshDF.replace({'Eco-Design Products': {'False': ''}}, inplace=True)
	addDataFrame(ws4, dnshDF)
	columnWidths(ws4, {'A': 12, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 14, 'I': 14, 'J': 14, 'K': 14, 'L': 14, 'M': 14, 'N': 14, 'O': 14, 'P': 14, 'Q': 14, 'R': 14, 'S': 14, 'T': 14, 'U': 14, 'V': 14, 'W': 14, 'X': 14, 'Y': 14, 'Z': 14, 'AA': 14, 'AB': 14, 'AC': 14, 'AD': 14, 'AE': 14, 'AF': 14, 'AG': 14, 'AH': 14, 'AI': 14, 'AJ': 14, 'AK': 14})
	ws4.freeze_panes = "B2"

	try:
		timestr = time.strftime("%Y%m%d-%H%M%S_")
		workbook.save(timestr + rFileName)
	except PermissionError:
		print('Error: Unable to write report file')



#==============================================
def main():
#==============================================
	print('--------------------------------')
	print('Portfolio - EU Taxonomy for Climate change calculator, version: 0.7')
	print('--------------------------------')
	
	# initialize
	print('Connecting to Eikon...')
	init(args.APP_KEY)
	print('Reading input portfolio')
	ricList = loadInputPortfolio(args.input)
	print('Portfolio contains [%s] instruments: %s ...' % (len(ricList), ricList[0:4]))
	
	# load the database
	loadDatabase('database.xlsx')
	
	print('Mapping database loaded')
	print('Getting Segment/ESG data for portfolio...')
	# get data from Refinitiv
	taxonMaster, esgMaster, dnshMaster = getData(ricList)
	
	print('Data received, calculating taxonomy ratios...')
	
	sectorDF = pd.DataFrame()
	orgDF = pd.DataFrame(columns = ['Instrument', 'Name', 'Delisted', 'ESG Score', 'Economic Sector', 'TRBC Activity', 'Aligned by Industry', 'Additional Testing Required', 'Eligible', 'Not In Scope', 'Others', 'Aligned- Pass', 'Aligned- No Data', 'Aligned- Not in Scope', 'Additional testing needed', 'Total', 'Parent Eligible', 'Parent Eligible ratio', 'Parent Not In Scope ratio'])
	
	# process taxo data for each instrument in the list
	for instr in ricList:
		# get sub frame for this instrument
		subDF = pd.DataFrame(taxonMaster[taxonMaster['Instrument'] == instr])
		subDF.reset_index(inplace=True, drop=True)
		esgData = pd.DataFrame(esgMaster[esgMaster['Instrument'] == instr])
		esgData.reset_index(inplace=True, drop=True)
		report, msubDF = getTaxoForRic(instr, subDF, esgData)
		# append data to the aggregate data frame
		sectorDF = sectorDF.append(msubDF, ignore_index=True)
		orgDF = orgDF.append(report, ignore_index=True)
	
	print('Generating report')
	generateReport(args.report, orgDF, sectorDF, dnshMaster)

	
#==============================================
if __name__ == "__main__":
#==============================================
	# read input arguments
	parser = ArgumentParser()
	parser.add_argument('APP_KEY', help='Eikon AppKey. See the install readme help on how to generate one')
	parser.add_argument('-i', '--input', default='input.xlsx', help='Excel portfolio file containing the list of securities with a \'RIC\' column-header')
	parser.add_argument('-r', '--report', default='report.xlsx', help='Output report excel filename')
	args = parser.parse_args()

	# start processing
	main()
	
